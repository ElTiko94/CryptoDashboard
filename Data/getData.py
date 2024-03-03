"""
An interactive tool designed to provide real-time insights into cryptocurrency portfolios. 
It leverages APIs from CoinMarketCap and Binance to fetch up-to-date information, 
offering users a comprehensive view of their digital assets.
"""
import json
from os import environ, path
import sys
import subprocess
import datetime
import openpyxl
from requests import Session
from requests.exceptions import Timeout, TooManyRedirects
import xlwings as xw
from binance_api import BinanceApi


def main():
    """
    main function
    """
    #today = datetime.datetime.now()
    today = datetime.datetime(2024, 1, 1)
    star_atlas_j0 = datetime.datetime(2021, 12, 17)

    crypto_path = environ.get('crypto_path')
    file_name = path.join(crypto_path, "Data/Historique d'achats.xlsx")

    with open(path.join(crypto_path, 'Data/config.json')) as config_file:
        config = json.load(config_file)

    # Create a session object
    session = Session()

    binance_api = BinanceApi(config['binance_api_key'], config['binance_api_secret'], session)

    rewards = binance_api.get_cumulative_total_rewards()
    auto_invest = binance_api.get_auto_invest(config['plan_id_to_name'])

    while True :
        # dict : token_prices[token]
        token_prices = get_crypto_prices(config['tokens'], config['coinmarketcap_api_key'], session)


        if token_prices is not None:
            star_atlas_days = int((today - star_atlas_j0).days)
            update_excel_file(file_name, token_prices, rewards, auto_invest, star_atlas_days)
            print("")

        try :
            loop = int( sys.argv[1] )

        except IndexError :
            loop = 0

        if loop == 0 :
            break
        if redo := input("\nPress any key to Redo, or press 'N' to stop ? ") :
            if redo.upper() == 'N':
                break

    # Path to your bash script
    batch_script = path.join(crypto_path, 'Data/Open_excel.bat')

    # Run the script
    subprocess.run([batch_script], shell=True)

    input("\nPress Enter to close the Excel workbook...")

def get_crypto_prices(symbols, api_key, session):
    """
    Get token price from token in symbols list, with Coinmarket cap API"
    """
    url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest'
    parameters = {'symbol': ','.join(symbols)}
    headers = {'Accepts': 'application/json', 'X-CMC_PRO_API_KEY': api_key}

    # Update the session headers for this request
    session.headers.update(headers)

    try:
        response = session.get(url, params=parameters)
        data = json.loads(response.text)
        if data.get('status', {}).get('error_code') != 0:
            print(f"Error in API response: {data.get('status', {}).get('error_message')}")
            return None
    except (ConnectionError, Timeout, TooManyRedirects) as e:
        print(e)
        return None

    prices = {}
    for symbol in symbols:
        # Ensure the symbol exists in the response data
        if symbol in data.get("data", {}):
            prices[symbol] = data["data"][symbol]["quote"]["USD"]["price"]
        else:
            print(f"Price data for {symbol} not found in response.")

    return prices

def update_excel_file(file_name, token_prices, rewards, auto_invest, days):
    """
    Update the excel history file, with prices up to date, staking rewards and DCA information
    """
    excel_file = openpyxl.load_workbook(file_name)

    # Update Star Atlas number of Stacking days
    cell = excel_file['ATLAS'].cell(row=2, column=8)  # H2
    cell.value = days

    app = xw.App(visible=False)
    book = app.books.open(file_name)

    for token, price in token_prices.items():
        sheet = excel_file[str(token)]
        cell = sheet.cell(row=3, column=10)  # J3
        # Update Tokens prices
        cell.value = float(price)

        print_valid_transactions(book.sheets[token], price, token)

        if token in rewards:
            cell = get_yellow_cells(sheet)
            sheet.cell(row=cell[0][0], column=cell[0][1]).value = float(rewards[token])

    book.close()
    app.quit()

    for plan in auto_invest:
        for detail in plan["details"]:
            sheet = excel_file[str(detail['targetAsset'])]
            cell_row = get_dca_cell(sheet, plan['planId'])
            sheet.cell(row=cell_row, column=2).value = float(detail['purchasedAmount'])
            sheet.cell(row=cell_row, column=4).value = float(detail['totalInvestedInUSD'])

    # Save changes
    excel_file.save(file_name)

def print_valid_transactions(sheet, token_price, token):
    """
    Look for transactions that could be done and printing it
    """
    row = 3
    price_cell_value = sheet.range(f'O{row}').value

    print(f'token : {token} price : {float(token_price)}$')
    if price_cell_value is not None and float(price_cell_value) > float(token_price):
        print_transaction("Buy", sheet.range(f'N{row}').value, token, sheet.range(f'P{row}').value)
        return None

    if token not in ['ETH','BTC'] :
        for row in range(6, 40):
            price_cell_value = sheet.range(f'O{row}').value

            if price_cell_value is not None and price_cell_value!= "Token Price" :
                status_cell_value = sheet.range(f'Q{row}').value
                if status_cell_value != "Done":
                    if float(price_cell_value) < float(token_price) and token != "SHIB":
                        print_transaction("Sell", sheet.range(f'N{row}').value, token, sheet.range(f'P{row}').value)
                        return None
                    if (token == "SHIB") :
                        if float(sheet.range(f'P{row}').value)/sheet.range(f'N{row}').value < float(token_price):
                            print_transaction("Sell", sheet.range(f'N{row}').value, token, sheet.range(f'P{row}').value)
                            return None
    return None

def get_dca_cell(sheet, plan_id, column_letter='E'):
    """
    Look for DCA cell correponding to the plan_id in a specific sheet
    """
    for row in sheet[column_letter]:
        if row.value == plan_id:
            return row.row
    return None

def get_yellow_cells(sheet, column_letter='B'):
    """
    Find filled yellow cell
    """
    yellow_cells = []

    # Iterate through each cell in the specified column
    for row in sheet[column_letter]:
        if row.fill.start_color.index == 'FFFFFF00':  # Yellow fill color in hex
            yellow_cells.append((row.row, row.column))
            break

    return yellow_cells

def print_transaction(transaction_type: str, amount: float, token: str, price: float):
    """
    Print a valid transaction template.

    Parameters:
        transaction_type (str): The type of transaction (e.g., 'Buy', 'Sell').
        amount (float): The amount of tokens involved in the transaction.
        token (str): The name or symbol of the token.
        price (float): The price of the transaction.

    Prints:
        A formatted string displaying transaction details.
    """
    print(f"     {transaction_type} {amount} of {token} for ${price:.2f}")

if __name__ == "__main__" :
    main()
